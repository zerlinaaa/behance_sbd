import re
import json
import openpyxl
import sys
from openpyxl.styles import Font, Alignment, PatternFill

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Load data
with open("behance_data.json", "r", encoding="utf-8") as f:
    data = json.load(f)


print("=" * 70)
print("HASIL PENGECEKAN DATA REDUNDAN (LOGIKA OR)")
print("Eliminasi jika title SAMA ATAU slug SAMA")
print("=" * 70)

seen_titles = {}      
seen_slugs  = {}     
redundan = []         

for i, item in enumerate(data):
    title_lower = item.get("title", "").strip().lower()
    slug_lower  = item.get("slug", "").strip().lower()

    title_match = title_lower in seen_titles
    slug_match  = slug_lower  in seen_slugs

    if title_match or slug_match:
        redundan.append({
            "index": i,
            "title": item.get("title"),
            "slug": item.get("slug"),
            "matched_on": [],
            "match_details": []
        })
        if title_match:
            redundan[-1]["matched_on"].append("title")
            redundan[-1]["match_details"].append(f"title='{item.get('title')}' → sama dengan data #{seen_titles[title_lower]+1}")
        if slug_match:
            redundan[-1]["matched_on"].append("slug")
            redundan[-1]["match_details"].append(f"slug='{slug_lower}' → sama dengan data #{seen_slugs[slug_lower]+1}")
    else:
        seen_titles[title_lower] = i
        seen_slugs[slug_lower]  = i

# Tampilkan hasil
if redundan:
    print(f"\nDitemukan {len(redundan)} data redundan:\n")
    for idx, r in enumerate(redundan, 1):
        print(f"  [{idx}] DATA #{r['index']+1}  —  ELIMINASI")
        print(f"      Title : {r['title']}")
        print(f"      Slug  : {r['slug']}")
        print(f"      Match via: {', '.join(r['matched_on']).upper()}")
        for detail in r['match_details']:
            print(f"        → {detail}")
        print()
else:
    print("\n(Tidak ada data redundan)")


clean_data = [item for i, item in enumerate(data)
              if i not in {r['index'] for r in redundan}]

print("=" * 70)
print(f"Total data original  : {len(data)}")
print(f"Data redundan dihapus: {len(redundan)}")
print(f"Sisa data bersih     : {len(clean_data)}")

print(f"\n{'='*70}")
print("MEMBUAT FILE EXCEL...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Behance Data"

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = ["id", "user_id", "title", "slug", "description", "cover_image",
           "status", "views_count", "likes_count", "comments_count"]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

col_widths = [8, 10, 40, 50, 80, 70, 12, 12, 12, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

def clean_text(text):
    if not text:
        return ""
    return re.sub(r'[\x00-\x1f\x7f-\x9f]', '', str(text))

for row_idx, item in enumerate(clean_data, 1):
    row_num = row_idx + 1
    ws.cell(row=row_num, column=1, value=f"BH-{row_idx:04d}")
    ws.cell(row=row_num, column=2, value=item.get("user_id", ""))
    clean_title = clean_text(item.get("title", ""))
    ws.cell(row=row_num, column=3, value=clean_title)
    ws.cell(row=row_num, column=4, value=item.get("slug", ""))
    ws.cell(row=row_num, column=5, value=item.get("description", ""))
    ws.cell(row=row_num, column=6, value=item.get("cover_image", ""))
    ws.cell(row=row_num, column=7, value=item.get("status", ""))
    ws.cell(row=row_num, column=8, value=item.get("views_count", 0))
    ws.cell(row=row_num, column=9, value=item.get("likes_count", 0))
    ws.cell(row=row_num, column=10, value=item.get("comments_count", 0))

    if row_idx % 2 == 0:
        for c in range(1, 11):
            ws.cell(row=row_num, column=c).fill = PatternFill(
                start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(clean_data)+1}"

wb.save("behance_cleaned.xlsx")

print(f"SUKSES! File tersimpan: behance_cleaned.xlsx")
print(f"Total baris data: {len(clean_data)} (+ 1 header row)")
print("ID format: BH-0001, BH-0002, ... BH-N")