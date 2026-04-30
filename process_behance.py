"""
============================================================
BEHANCE DATA PROCESSOR — All-in-One
============================================================
Urutan proses:
  1. Validasi data (cek kosong / incomplete)
  2. Convert ke CSV
  3. Convert ke Excel (.xlsx)
  4. Buat SQL dump
  5. Buat Laravel Seeder (PHP)
============================================================
Cara pakai:
  python process_behance.py

Pastikan behance_data.json ada di folder yang sama.
============================================================
"""

import json, csv, os, re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠  openpyxl tidak ditemukan. Install: pip install openpyxl")

# ── Config ──────────────────────────────────────────────────────────────────
INPUT_FILE   = "behance_data.json"
OUT_CSV      = "behance_data.csv"
OUT_EXCEL    = "behance_data.xlsx"
OUT_SQL      = "behance_data.sql"
OUT_SEEDER   = "BehanceProjectsSeeder.php"
TABLE_NAME   = "projects"

REQUIRED_FIELDS = ["title", "slug", "status"]
NUMERIC_FIELDS  = ["views_count", "likes_count", "comments_count", "user_id"]

# ── Warna header Excel ───────────────────────────────────────────────────────
HEADER_BG   = "2563EB"   # biru
HEADER_FG   = "FFFFFF"   # putih
ALT_ROW_BG  = "EFF6FF"   # biru muda alt row


# =============================================================================
# 1. LOAD & VALIDASI
# =============================================================================
def load_and_validate(path: str):
    print("\n" + "="*55)
    print("STEP 1 — VALIDASI DATA")
    print("="*55)

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    total   = len(raw)
    clean   = []
    issues  = []

    for i, item in enumerate(raw):
        row_issues = []

        # Pastikan semua field wajib ada & tidak kosong
        for field in REQUIRED_FIELDS:
            if not item.get(field, "").strip():
                row_issues.append(f"'{field}' kosong")

        # Pastikan field numerik valid
        for field in NUMERIC_FIELDS:
            val = item.get(field, 0)
            try:
                item[field] = int(val)
            except (ValueError, TypeError):
                item[field] = 0
                row_issues.append(f"'{field}' bukan angka → di-set 0")

        # Default value untuk field opsional (lakukan SEBELUM cek slug)
        item.setdefault("description", "")
        item.setdefault("cover_image", "")
        item.setdefault("status", "published")

        # Buat slug jika kosong/tidak ada
        if not item.get("slug"):
            import uuid
            safe = re.sub(r"[^a-zA-Z0-9\s]", "", item.get("title", "project"))
            base = safe.strip().lower().replace(" ", "-")[:60]
            item["slug"] = f"{base}-{uuid.uuid4().hex[:8]}"
            row_issues.append("'slug' kosong → di-generate otomatis")

        # Pastikan slug tidak duplikat
        item["slug"] = item.get("slug") or f"project-{random.randint(10000,99999)}"

        slugs_so_far = [c.get("slug") for c in clean if c.get("slug")]

        if item["slug"] in slugs_so_far:
            item["slug"] = f"{item['slug']}-{i}"
            row_issues.append("slug duplikat → ditambah suffix")

        if row_issues:
            issues.append({"row": i + 1, "title": item.get("title", "?"), "issues": row_issues})

        clean.append(item)

    # Laporan
    print(f"  Total data    : {total}")
    print(f"  Data bersih   : {len(clean)}")
    print(f"  Baris bermasalah: {len(issues)}")

    if issues:
        print("\n  Detail masalah (max 10 pertama):")
        for issue in issues[:10]:
            print(f"    Baris {issue['row']:>4} [{issue['title'][:30]}]: {', '.join(issue['issues'])}")
        if len(issues) > 10:
            print(f"    ... dan {len(issues)-10} baris lainnya")

    print(f"\n  ✅ Validasi selesai. {len(clean)} data siap diproses.")
    return clean


# =============================================================================
# 2. EXPORT CSV
# =============================================================================
def export_csv(data: list, path: str):
    print("\n" + "="*55)
    print("STEP 2 — EXPORT CSV")
    print("="*55)

    fields = ["user_id","title","description","cover_image",
              "slug","status","views_count","likes_count","comments_count"]

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)

    size = os.path.getsize(path) / 1024
    print(f"  ✅ Tersimpan: {path} ({size:.1f} KB, {len(data)} baris)")


# =============================================================================
# 3. EXPORT EXCEL
# =============================================================================
def export_excel(data: list, path: str):
    print("\n" + "="*55)
    print("STEP 3 — EXPORT EXCEL")
    print("="*55)

    if not HAS_OPENPYXL:
        print("  ⚠ Dilewati (openpyxl tidak terinstall)")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    headers = ["user_id","title","description","cover_image",
               "slug","status","views_count","likes_count","comments_count"]

    header_font  = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
    header_fill  = PatternFill("solid", start_color=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")
    alt_fill     = PatternFill("solid", start_color=ALT_ROW_BG)
    thin_border  = Border(
        left   = Side(style="thin", color="CCCCCC"),
        right  = Side(style="thin", color="CCCCCC"),
        top    = Side(style="thin", color="CCCCCC"),
        bottom = Side(style="thin", color="CCCCCC"),
    )

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.upper().replace("_", " "))
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_i, item in enumerate(data, 2):
        fill = alt_fill if row_i % 2 == 0 else None
        for col_i, key in enumerate(headers, 1):
            val  = item.get(key, "")
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font   = Font(name="Arial", size=9)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            # Angka rata kanan
            if key in NUMERIC_FIELDS:
                cell.alignment = Alignment(horizontal="right")

    # Lebar kolom otomatis
    col_widths = {
        "USER ID": 8, "TITLE": 35, "DESCRIPTION": 40, "COVER IMAGE": 50,
        "SLUG": 35, "STATUS": 10, "VIEWS COUNT": 12,
        "LIKES COUNT": 12, "COMMENTS COUNT": 14,
    }
    for col_i, h in enumerate(headers, 1):
        key = h.upper().replace("_", " ")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = col_widths.get(key, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    summary_data = [
        ("Total Projects",    len(data)),
        ("Published",         sum(1 for d in data if d.get("status") == "published")),
        ("Total Views",       f'=SUM(Data!G2:G{len(data)+1})'),
        ("Total Likes",       f'=SUM(Data!H2:H{len(data)+1})'),
        ("Total Comments",    f'=SUM(Data!I2:I{len(data)+1})'),
        ("Avg Views",         f'=AVERAGE(Data!G2:G{len(data)+1})'),
        ("Generated At",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    ws2.cell(1, 1, "METRIC").font  = Font(bold=True, color=HEADER_FG, name="Arial")
    ws2.cell(1, 1).fill            = PatternFill("solid", start_color=HEADER_BG)
    ws2.cell(1, 2, "VALUE").font   = Font(bold=True, color=HEADER_FG, name="Arial")
    ws2.cell(1, 2).fill            = PatternFill("solid", start_color=HEADER_BG)

    for r, (metric, value) in enumerate(summary_data, 2):
        ws2.cell(r, 1, metric).font = Font(name="Arial", size=10)
        cell = ws2.cell(r, 2, value)
        cell.font = Font(name="Arial", size=10, bold=True)
        if isinstance(value, (int, float)):
            cell.alignment = Alignment(horizontal="right")

    wb.save(path)
    size = os.path.getsize(path) / 1024
    print(f"  ✅ Tersimpan: {path} ({size:.1f} KB, {len(data)} baris + sheet Summary)")


# =============================================================================
# 4. EXPORT SQL
# =============================================================================
def export_sql(data: list, path: str):
    print("\n" + "="*55)
    print("STEP 4 — EXPORT SQL DUMP")
    print("="*55)

    def esc(val):
        if val is None:
            return "NULL"
        if isinstance(val, int):
            return str(val)
        # Escape single quotes
        return "'" + str(val).replace("\\", "\\\\").replace("'", "\\'") + "'"

    lines = []
    lines.append(f"-- Behance Projects SQL Dump")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Total: {len(data)} rows")
    lines.append("")
    lines.append(f"SET NAMES utf8mb4;")
    lines.append(f"SET FOREIGN_KEY_CHECKS = 0;")
    lines.append("")
    lines.append(f"-- Table structure")
    lines.append(f"CREATE TABLE IF NOT EXISTS `{TABLE_NAME}` (")
    lines.append(f"  `id` bigint(20) UNSIGNED NOT NULL AUTO_INCREMENT,")
    lines.append(f"  `user_id` bigint(20) UNSIGNED NOT NULL DEFAULT 1,")
    lines.append(f"  `title` varchar(255) NOT NULL,")
    lines.append(f"  `description` text DEFAULT NULL,")
    lines.append(f"  `cover_image` text DEFAULT NULL,")
    lines.append(f"  `slug` varchar(255) NOT NULL,")
    lines.append(f"  `status` varchar(50) NOT NULL DEFAULT 'published',")
    lines.append(f"  `views_count` int(11) NOT NULL DEFAULT 0,")
    lines.append(f"  `likes_count` int(11) NOT NULL DEFAULT 0,")
    lines.append(f"  `comments_count` int(11) NOT NULL DEFAULT 0,")
    lines.append(f"  `created_at` timestamp NULL DEFAULT NULL,")
    lines.append(f"  `updated_at` timestamp NULL DEFAULT NULL,")
    lines.append(f"  PRIMARY KEY (`id`),")
    lines.append(f"  UNIQUE KEY `projects_slug_unique` (`slug`)")
    lines.append(f") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;")
    lines.append("")
    lines.append(f"-- Data")

    # Insert dalam batch 100
    BATCH = 100
    now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cols  = "(user_id, title, description, cover_image, slug, status, views_count, likes_count, comments_count, created_at, updated_at)"

    for start in range(0, len(data), BATCH):
        batch = data[start:start + BATCH]
        lines.append(f"INSERT INTO `{TABLE_NAME}` {cols} VALUES")
        rows = []
        for item in batch:
            rows.append(
                f"  ({esc(item['user_id'])}, {esc(item['title'])}, "
                f"{esc(item['description'])}, {esc(item['cover_image'])}, "
                f"{esc(item['slug'])}, {esc(item['status'])}, "
                f"{esc(item['views_count'])}, {esc(item['likes_count'])}, "
                f"{esc(item['comments_count'])}, {esc(now)}, {esc(now)})"
            )
        lines.append(",\n".join(rows) + ";")
        lines.append("")

    lines.append("SET FOREIGN_KEY_CHECKS = 1;")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    size = os.path.getsize(path) / 1024
    print(f"  ✅ Tersimpan: {path} ({size:.1f} KB, {len(data)} rows dalam batch {BATCH})")


# =============================================================================
# 5. EXPORT LARAVEL SEEDER
# =============================================================================
def export_seeder(data: list, path: str):
    print("\n" + "="*55)
    print("STEP 5 — EXPORT LARAVEL SEEDER")
    print("="*55)

    def php_str(val):
        if val is None or val == "":
            return "null"
        return "'" + str(val).replace("\\", "\\\\").replace("'", "\\'") + "'"

    def php_int(val):
        try:
            return str(int(val))
        except:
            return "0"

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("<?php")
    lines.append("")
    lines.append("namespace Database\\Seeders;")
    lines.append("")
    lines.append("use Illuminate\\Database\\Seeder;")
    lines.append("use Illuminate\\Support\\Facades\\DB;")
    lines.append("")
    lines.append("class BehanceProjectsSeeder extends Seeder")
    lines.append("{")
    lines.append("    public function run(): void")
    lines.append("    {")
    lines.append(f"        // Total: {len(data)} projects")
    lines.append(f"        // Generated: {now}")
    lines.append("")
    lines.append("        $now = now();")
    lines.append("")
    lines.append("        $data = [")

    for item in data:
        lines.append("            [")
        lines.append(f"                'user_id'        => {php_int(item['user_id'])},")
        lines.append(f"                'title'          => {php_str(item['title'])},")
        lines.append(f"                'description'    => {php_str(item['description'])},")
        lines.append(f"                'cover_image'    => {php_str(item['cover_image'])},")
        lines.append(f"                'slug'           => {php_str(item['slug'])},")
        lines.append(f"                'status'         => {php_str(item['status'])},")
        lines.append(f"                'views_count'    => {php_int(item['views_count'])},")
        lines.append(f"                'likes_count'    => {php_int(item['likes_count'])},")
        lines.append(f"                'comments_count' => {php_int(item['comments_count'])},")
        lines.append(f"                'created_at'     => $now,")
        lines.append(f"                'updated_at'     => $now,")
        lines.append("            ],")

    lines.append("        ];")
    lines.append("")
    lines.append("        // Insert dalam chunk 100 supaya tidak timeout")
    lines.append("        foreach (array_chunk($data, 100) as $chunk) {")
    lines.append(f"            DB::table('{TABLE_NAME}')->insert($chunk);")
    lines.append("        }")
    lines.append("    }")
    lines.append("}")
    lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    size = os.path.getsize(path) / 1024
    print(f"  ✅ Tersimpan: {path} ({size:.1f} KB)")
    print(f"\n  Cara pakai di Laravel:")
    print(f"  1. Copy {path} → database/seeders/")
    print(f"  2. Daftarkan di DatabaseSeeder.php:")
    print(f"       $this->call(BehanceProjectsSeeder::class);")
    print(f"  3. Jalankan: php artisan db:seed --class=BehanceProjectsSeeder")


# =============================================================================
# MAIN
# =============================================================================
def main():
    print("\n" + "="*55)
    print(" BEHANCE DATA PROCESSOR — ALL IN ONE")
    print("="*55)

    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ File '{INPUT_FILE}' tidak ditemukan!")
        print("   Pastikan kamu sudah menjalankan behance_scraper.py terlebih dahulu.")
        return

    data = load_and_validate(INPUT_FILE)

    export_csv(data, OUT_CSV)
    export_excel(data, OUT_EXCEL)
    export_sql(data, OUT_SQL)
    export_seeder(data, OUT_SEEDER)

    print("\n" + "="*55)
    print(" SEMUA PROSES SELESAI ✅")
    print("="*55)
    print(f"  📄 {OUT_CSV}")
    print(f"  📊 {OUT_EXCEL}")
    print(f"  🗄  {OUT_SQL}")
    print(f"  🐘 {OUT_SEEDER}")
    print("="*55 + "\n")


if __name__ == "__main__":
    main()